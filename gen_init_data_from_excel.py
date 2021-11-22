import openpyxl
import json 

def gen(src_file, dst_file):
    aff = {
        "france": 4,
        "spain": 2,
        "chili": 2,
        "australia": 2
    }
    wb = openpyxl.load_workbook(src_file)
    sh = wb["MJ_Initialisation"]
    h = 24
    

    # get pv
    pv = {a: {p: [None for _ in range(h)] for p in [f"P{k+1}" for k in range(np)]} for a, np in aff.items()}
    offset = 6
    for a, np in aff.items():
        for i in range(np):
            for t in range(h):
                pv[a][f"P{i+1}"][t] = int(sh.cell(offset, 4+t).value)
            offset+=1
    
    # get pa 
    offset = 20
    pa = {a: {p: [None for _ in range(h)] for p in [f"P{k+1}" for k in range(np)]} for a, np in aff.items()}
    for a, np in aff.items():
        for i in range(np):
            for t in range(h):
                pa[a][f"P{i+1}"][t] = int(sh.cell(offset, 4+t).value)
            offset+=2

    # get pdp
    offset = 44
    pdp = {p: [None for _ in range(h)] for p in ["P1", "P2", "P3", "P4"]}
    for i in range(4):
        for t in range(h):
            pdp[f"P{i+1}"][t] = int(sh.cell(offset, 4+t).value)
        offset+=2
    with open(dst_file) as fp:
        data = json.load(fp)

    data["sales_forcast"] = pv
    data["prev_supply_plan"] = pa
    data["prev_prod_plan"] = pdp

    with open(dst_file, "w") as fp:
        json.dump(data, fp)
 




if __name__=="__main__":
    gen("S2-Monthly-firstweek-correct-formula-VF.xlsm", "simu_inputs/input_S2.json")