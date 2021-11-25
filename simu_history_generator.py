import openpyxl
import json 
import os

def sumOverAffiliates(quantity, affiliates, products, horizon):
    return {
            p: [
                sum([quantity[a][p][t] for a in affiliates if p in quantity[a]]) for t in range(horizon)
            ] for p in products
        }
    
def run():
    with open(f"simu_inputs/global_input.json") as fp:
        data = json.load(fp)
    products = data["products"]
    affiliates = data["affiliates"]
    horizon = data["horizon"]
    
    wb = openpyxl.load_workbook("templates/template_simu_result.xlsx")
    
    product_sales_history = []
    prod_plan_history = []
    supply_demand_history = []
    supply_plan_history = []
    prod_demand_history = []
    
    for file_name in os.listdir("simu_inputs"):
        if file_name.startswith("input_S"):
            with open(f"simu_inputs/{file_name}") as fp:
                input_data = json.load(fp)
            product_sales_history.append(sumOverAffiliates(input_data["sales_forcast"], affiliates, products, horizon))
            
    for file_name in os.listdir("simu_history"):
        if file_name.startswith("snapshot_S"):
            with open(f"simu_history/{file_name}") as fp:
                snapshot = json.load(fp)
            supply_demand_history.append(sumOverAffiliates(snapshot["supply_demand"], affiliates, products, horizon))
            prod_plan_history.append(snapshot["prod_plan"])
            supply_plan_history.append(sumOverAffiliates(snapshot["supply_plan"], affiliates, products, horizon))
            prod_demand_history.append(snapshot["prod_demand"])
    
    nbr_weeks = len(prod_plan_history)
    
    for p in products:
        for w in range(nbr_weeks):
            for t in range(horizon):
                wb["PV"].cell(row=3+w, column=3+t+w).value = product_sales_history[w][p][t]
                wb["BA"].cell(row=3+w, column=3+t+w).value = supply_demand_history[w][p][t]
                wb["BP"].cell(row=3+w, column=3+t+w).value = prod_demand_history[w][p][t]
                wb["PDP"].cell(row=3+w, column=3+t+w).value = prod_plan_history[w][p][t]
                wb["PA"].cell(row=3+w, column=3+t+w).value = supply_plan_history[w][p][t]
        if not os.path.exists("simu_excel_results"):
            os.mkdir("simu_excel_results")
        wb.save(f"simu_excel_results/{p}_results.xlsx")