import openpyxl
import json 
import os

def sumOverAffiliates(quantity, affiliates, products, horizon):
    return {
            p: [
                sum([quantity[a][p][t] for a in affiliates if p in quantity[a]])
                for t in range(horizon)
            ] for p in products
        }
    
def writeToExcel():
    with open(f"simu_inputs/input_S2.json") as fp:
        data = json.load(fp)
    products = data["products"]
    affiliates = data["affiliates"]
    horizon = data["horizon"]
    wb = openpyxl.load_workbook("templates/template_simu_result.xlsx")
    
    product_sales = []
    prod_plan = []
    supply_demand = []
    supply_plan = []
    prod_demand = []
    
    for file_name in os.listdir("simu_inputs"):
        if "input_S" in file_name:
            with open(f"simu_inputs/{file_name}") as fp:
                input_data = json.load(fp)
            product_sales.append(sumOverAffiliates(input_data["sales_forcast"], affiliates, products, horizon))
            
    for file_name in os.listdir("simu_outputs"):
        with open(f"simu_outputs/{file_name}") as fp:
            output_data = json.load(fp)
        supply_demand.append(sumOverAffiliates(output_data["supply_demand"], affiliates, products, horizon))
        prod_plan.append(output_data["prod_plan"])
        supply_plan.append(sumOverAffiliates(output_data["supply_plan"], affiliates, products, horizon))
        prod_demand.append(output_data["prod_demand"])
    
    nbr_weeks = len(prod_plan)
    for p in products:
        for w in range(nbr_weeks):
            for t in range(horizon):
                wb["PV"].cell(row=3+w, column=3+t+w).value = product_sales[w][p][t]
                wb["BA"].cell(row=3+w, column=3+t+w).value = supply_demand[w][p][t]
                wb["BP"].cell(row=3+w, column=3+t+w).value = prod_demand[w][p][t]
                wb["PDP"].cell(row=3+w, column=3+t+w).value = prod_plan[w][p][t]
                wb["PA"].cell(row=3+w, column=3+t+w).value = supply_plan[w][p][t]
        wb.save(f"simu_excel_results/{p}_results.xlsx")

if __name__ == "__main__":
    writeToExcel()