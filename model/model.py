import json
import copy
import logging
import openpyxl
from . import Shared, Affiliate, CDC, Factory, utils

class Model(Shared):
    
    def __init__(self) -> None:
        super().__init__()     
        self.week = None
        self.initial_stock = None
        self.prev_production = None
        self.prev_supply = None
        self.sales_forcast = None
        self.affiliates = {name: Affiliate(name) for name in self.itAffiliates()}
        self.factory = Factory()
        self.cdc = CDC()

    def getCDCProductSalesForcast(self):
        cdc_sales = {a: 
            {p: self.sales_forcast[a][p][int(aff["delivery_time"]):] + [0] * int(aff["delivery_time"]) for p in self.itAffProducts(a)}     
            for a, aff in self.settings["affiliate"].items()
        }
        return self.sumOverAffiliate(cdc_sales)
    
    def loadSalesForcast(self, file_p):
        with open(file_p) as fp:
            self.sales_forcast = json.load(fp)

    def loadWeekInput(self, input_file=None, input_dict=None):
        if input_file:
            with open(input_file) as json_file:
                input_dict = json.load(json_file)
        if input_dict is None:
            raise Exception("No input file or dict given!")
        self.week: int = input_dict["week"]
        self.initial_stock: dict[str: int] = input_dict["initial_stock"]
        self.prev_production: dict[str, list[int]] = input_dict["prev_production"]
        self.prev_supply: dict[str, dict[str, list[int]]] = input_dict["prev_supply"]

    def getAffiliateSupplyDemand(self):
        return {name: a.demand for name, a in self.affiliates.items()}
    
    def getNextInitialStock(self):
        initial_stock = {}
        for a, aff in self.affiliates.items():
            initial_stock[a] = {
                p: aff.initial_stock[p] + self.prev_supply[a][p][0] +\
                    (self.cdc.supply[a][p][0] if aff.delivery_time==0 else 0) - self.sales_forcast[a][p][0] 
                for p in aff.products
            } 
        initial_stock["cdc"] = {p: self.cdc.projected_stock[p][0] for p in self.products}
        return initial_stock

    def getNextSupply(self):
        next_supply = {}
        for a, aff in self.affiliates.items():
            next_supply[a] = {}
            for p in self.itAffProducts(a):
                next_supply[a][p] = self.prev_supply[a][p][1:aff.delivery_time+1] +\
                    self.cdc_supply[a][p][1:self.horizon-aff.delivery_time] +\
                        [self.cdc_supply[a][p][self.horizon-aff.delivery_time-1]]
        return next_supply

    def getNextProdPlan(self):
        production = {}
        for p in self.products:
            production[p] = self.prev_production[p][1:self.prod_time+1] + self.factory.production[p][1:self.horizon-self.prod_time]
            production[p] += [production[p][-1]]
        return production

    def generateNextWeekInput(self, file_path):
        data = {}
        data["prev_supply"] = self.getNextSupply()
        data["prev_production"] = self.getNextProdPlan()
        data["initial_stock"] = self.getNextInitialStock()
        data["week"] = self.week + 1
        with open(file_path, 'w') as fp:
            json.dump(data, fp)
        return data

    def getAffiliatesDemand(self, prev_supply, sales_forcast, w):        
        aff_d = {}
        for a, affiliate in self.affiliates.items():
            affiliate.initial_stock = self.initial_stock[a]
            logging.debug(f"Calculation BA for {a}, week {w}")
            aff_d[a] = affiliate.getDemand(sales_forcast[a], prev_supply[a])
        return aff_d
            
    def runWeek(self, sales_forcast):
        self.sales_forcast = sales_forcast
        self.aff_demand = self.getAffiliatesDemand(self.sales_forcast, self.prev_supply, self.week)
        self.cdc_demand = self.getCDCDemand(self.aff_demand)
        self.cdc.initial_stock = self.getCDCInitialStock()
        self.cdc_prev_supply = self.getCDCPrevSupply()
        self.cdc_product_demand = self.sumOverAffiliate(self.cdc_demand)
        self.cdc_prod_demand = self.cdc.getProdDemand(self.prev_production, self.cdc_product_demand)
        self.factory_prod_demand = {p: self.cdc_prod_demand[p][self.prod_time:] + [0] * self.prod_time for p in self.products}
        self.factory_prev_production = {p: self.prev_production[p][self.prod_time:] + [0] * self.prod_time for p in self.products}
        self.factory.getProduction(self.factory_prod_demand, self.factory_prev_production)
        self.cdc_reception = self.getCDCReception()
        self.cdc_supply, self.cdc_product_supply, self.cdc_dept = self.cdc.getAffSupply(self.cdc_prev_supply, self.cdc_demand, self.cdc_reception)
 
    def getSnapShot(self):
        snap = {
            "week": self.week,
            "product_supply": copy.deepcopy(self.cdc_product_supply),
            "prod_demand": self.cdc_prod_demand,
            "sales_forcast": self.sales_forcast,
            "dept": self.cdc_dept,
            "initial_stock": self.getCDCInitialStock(),
            "metrics": {}
        }
        return snap
        
    def saveCDCSupply(self, file_path):
        with open(file_path, 'w') as fp:
            json.dump(self.cdc.supply, fp)

    def exportDataToExcel(self, dst_f):
        wb = openpyxl.load_workbook(self.platform_template_f)
        sheet = wb.active
        sheet.cell(2, 2).value = f"W{self.week}/20"
        cdc_demand = self.getCDCDemand()
        cdc_reception = self.getCDCReception()
        cdc_initial_stock = self.getCDCInitialStock()
        cdc_prev_supply = self.getCDCAffSupply()
        offset = 0
        # header week before
        sheet.cell(4, 8).value = f"W{self.week-1}/20"
        for t in range(self.horizon):
            # header weeks
            sheet.cell(4, 9 + t).value = f"W{self.week+t}/20"
        for p in self.products:
            product_block_start_row = 5 + offset
            # stock initial 
            sheet.cell(product_block_start_row + 1, 8).value = cdc_initial_stock[p]
            for t in range(self.horizon):
                # programmed_reception Factory -> CDC (pdp + queued)
                sheet.cell(product_block_start_row, 9 + t).value = cdc_reception[p][t]
                j = 0
                for a in self.affiliates.values():
                    if p in a.products:
                        # BA Affiliate -> CDC
                        sheet.cell(product_block_start_row + 2 + j * 2, 9 + t).value = cdc_demand[a.name][p][t]
                        # PA CDC -> affiliate 
                        sheet.cell(product_block_start_row + 2 + j * 2 + 1, 9 + t).value = cdc_prev_supply[a.name][p][t]
                        j += 1
            nbr_ff_p = sum([1 for aff in self.affiliates.values() if p in aff.products])
            offset += (2 * nbr_ff_p + 3)
        wb.save(dst_f)

    def loadDemandFromExcel(self, srf_f):
        demand = {}
        wb = openpyxl.load_workbook(srf_f)
        sheet = wb.active
        i = 0
        while sheet.cell(2 + i, 1).value:
            quantity = int(sheet.cell(2 + i, 2).value)
            affiliate = self.getAffByCode(sheet.cell(2 + i, 3).value)
            week = int(sheet.cell(2 + i, 6).value.split("/")[0][1:])
            product = sheet.cell(2 + i, 12).value
            if affiliate not in demand:
                demand[affiliate] = {}
            if product not in demand[affiliate]:
                demand[affiliate][product] = [None for _ in range(self.horizon)]
            demand[affiliate][product][week-self.week] = quantity
            i += 1
        return demand
    
    def getCDCDemand(self, aff_demand) -> dict[str, dict[str, list[int]]]:
        return {a: {
            p: aff_demand[a][p][aff.delivery_time:] + [0] * aff.delivery_time for p in aff.products
        } for a, aff in self.affiliates.items()}
     
    def getCDCInitialStock(self):
        return self.initial_stock["cdc"]
        
    def setCDCSupply(self, supply, product_supply):
        self.cdc_supply = supply
        self.cdc_product_supply = product_supply

    def getCDCReception(self):
        if self.week % 4 == 1:
            return {p: self.prev_production[p][:self.prod_time] + self.factory.production[p][:self.horizon-self.prod_time] for p in self.products}
        else:
            return self.prev_production

    def getCDCPrevSupply(self, prev_supply=None):
        if not prev_supply:
            prev_supply = self.prev_supply
        cdc_prev_supply = {a: 
            {p: prev_supply[a][p][aff.delivery_time:] + [0] * (self.horizon-aff.delivery_time) for p in self.itAffProducts(a)}     
            for a, aff in self.affiliates.items()
        }
        return cdc_prev_supply
    
