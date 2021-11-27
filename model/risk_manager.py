from random import randint
import re
import openpyxl
from model.model import Model
import model.utils as utils


class RiskManager:

    def __init__(self, horizon) -> None:
        self.horizon = horizon

    def loadDModel(self, model: Model, file_name: str):
        self.d_aff_model = {a: {p: {} for p in aff_products} for a, aff_products in model.affiliate_product.items()}
        self.d_model = {p: {} for p in model.products}
        wb = openpyxl.load_workbook(file_name)
        sh = wb.active
        r = 2
        while sh.cell(r, 1).value:
            product = sh.cell(r, 1).value
            aff_code = sh.cell(r, 2).value
            aff = model.affiliate_code[aff_code]
            param = sh.cell(r, 4).value
            if param == "RefWeek":
                quantity = utils.readRefWeekRow(sh, r, 5, self.horizon)
            else:
                quantity = utils.readSubRow(sh, r, 5, self.horizon)
            self.d_aff_model[aff][product][param] = quantity
            r += 1
    
    def getDpm(self, d, p):
        params = ["a", "b", "c", "d"]
        dist = {param: [0] * self.horizon for param in params + ["ref_week", "model_type"]}
        for a in d:
            if p in d[a]:
                Q = list(utils.accumu(d[a][p]))
                for t in range(self.horizon):
                    rw_t = self.d_aff_model[a][p]["RefWeek"][t]
                    model_type = self.d_aff_model[a][p]["ModelType"][t] 
                    for param in params:
                        alpha_t = self.d_aff_model[a][p][param][t]
                        F_t = Q[t] - Q[rw_t-2] if rw_t-2>0 else Q[t]
                        if model_type == "I1":
                            F_t /= t - (rw_t-1) + 1
                        dist[param][t] += Q[t] + alpha_t * F_t
        return dist
        
    def loadRModel(self, model: Model, file_name: str) -> None:
        wb = openpyxl.load_workbook(file_name)
        sh = wb.active
        params = ["a", "b", "c", "d", "model_type", "ref_week"]
        self.r_model = {p: {
            param: utils.readSubRow(sh, 2 + j + len(params) * k, 5, self.horizon) if param != "ref_week" else 
            utils.readRefWeekRow(sh, 2 + j + len(params) * k, 5, self.horizon) 
            for j, param in enumerate(params)
        } for k, p in enumerate(model.products) }
    
    def getRpm(self, r:dict, p: str) -> dict:
        params = ["a", "b", "c", "d"]
        Q = list(utils.accumu(r))
        dist = {param: [None] * self.horizon for param in params}
        for t in range(self.horizon):
            rw_t = self.r_model[p]["ref_week"][t]
            model_type = self.r_model[p]["model_type"][t]
            F_t = Q[t] - Q[rw_t-2] if rw_t-2>0 else Q[t]
            for param in params:
                alpha_t = self.r_model[p][param][t]
                if model_type == "I2":
                    dist[param][t] = round(Q[t] + alpha_t * F_t)
                elif model_type == "I1":
                    dist[param][t] = round(Q[t] + alpha_t * F_t / (t - (rw_t-1) + 1))
        return dist
    
    def l1p(self, a, b, s0, x):
        return utils.affineY(a + s0, b + s0, x)
    
    @staticmethod
    def _l4n(c, d, a, b, x):
        if a > d:
            return 1
        if x == c or x == b:
            return 0
        elif c >= b:
            return utils.affineY(c, d, x) + 1 - utils.affineY(a, b, x)
        elif c < b:
            x_star = ((b - a) * c + b * (d - c)) / (b - a + d - c)
            if x <= x_star :
                return 1 - utils.affineY(a, b, x)
            elif x > x_star:
                return utils.affineY(c, d, x)

    def l4n(self, rpm, dpm, s0, x, t):
        c, d = rpm["c"][t] + s0, rpm["d"][t] + s0 
        a, b = dpm["a"][t], dpm["b"][t]
        return self._l4n(c, d, a, b, x[t])

    def _l2p(self, c, d, x):
        return 1 - utils.affineY(d, c, x)

    def getL1Possibility(self, rpm: dict, x: dict, s0: dict) -> dict:
        l1_possibility = [self.l1p(rpm["a"][t], rpm["b"][t], s0, x[t]) for t in range(self.horizon)]
        return l1_possibility

    def getL2Possibility(self, dpm: dict, x: dict) -> dict:
        l2_possibility = [self._l2p(dpm["c"][t], dpm["d"][t], x[t]) for t in range(self.horizon)]
        return l2_possibility

    def getL4Possibility(self, l1p: dict, l2p: dict) -> dict:
        l4_possibility = [max(l1p[t], l2p[t]) for t in range(self.horizon)]
        return l4_possibility
    
    def getL4Necessity(self, rpm: dict, dpm:dict, x: dict, s0: dict) -> dict:
        l4_necessity = [self.l4n(rpm, dpm, s0, x, t) for t in range(self.horizon)]
        return l4_necessity

    def getG(self, rpm, dpm, s0, x):
        nl4p = self.getL4Necessity(rpm, dpm, x, s0)
        return max(nl4p)


if __name__ == "__main__":
    pass
