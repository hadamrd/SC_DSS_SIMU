
import openpyxl
from . import History, utils
import math


def getMean(q):
    size = len(q)
    return sum(q) / size

def var(q, m):
    size = len(q)
    sigma = sum([(q[t] - m)**2 for t in range(size)]) / size
    var = math.sqrt(sigma)
    return var

def getSigma(q, m):
    size = len(q)
    sigma = sum([(q[t] - m)**2 for t in range(size)]) / size
    return sigma

def getDiffHist(cq_hist, p, q_size, fh):
    res = [[None] * q_size] * (q_size - 1 - fh)
    for t in range(q_size - 1, 2 * q_size - 1):
        for k in range(fh-1, q_size - 2):
            y = t - k
            res[k - fh + 1][t - (q_size - 1)] = cq_hist[y][p][k] - cq_hist[y-1][p][k+1]
    return res

def getMeanVarDiffHist(cqdh):
    nrows = len(cqdh)
    mean_ = 0
    for w in range(nrows):
        mean_ += getMean(cqdh[w])
    mean_ /= nrows
    sigma_ = 0
    for w in range(nrows):
        sigma_ += getSigma(cqdh[w], mean_)
    var = math.sqrt(sigma_ / nrows)
    return mean_, var
        
def exportToExcel(hist1: History, hist2: History, dst_f):
    wb = openpyxl.load_workbook(hist1.indicators_template_f)
    sh = wb.active
    curr_row = 3
    col = 2
    nbr_weeks = hist1.nbr_weeks
    products = hist1.products
    for w in range(nbr_weeks):
        for p in products:
            sh.cell(curr_row, col + 1).value = hist2.metrics[w]["in"]["robustness"][p]
            sh.cell(curr_row, col + 2).value = hist1.metrics[w]["in"]["robustness"][p]
            sh.cell(curr_row, col + 3).value = hist1.metrics[w]["out"]["robustness"][p]
            
            sh.cell(curr_row, col + 6).value = hist2.metrics[w]["in"]["severity"][p]
            sh.cell(curr_row, col + 7).value = hist1.metrics[w]["in"]["severity"][p]
            sh.cell(curr_row, col + 8).value = hist1.metrics[w]["out"]["severity"][p]

            sh.cell(curr_row, col + 11).value = hist2.metrics[w]["in"]["frequency"][p]
            sh.cell(curr_row, col + 12).value = hist1.metrics[w]["in"]["frequency"][p]
            sh.cell(curr_row, col + 13).value = hist1.metrics[w]["out"]["frequency"][p]
            
            sh.cell(curr_row, col + 16).value = hist2.metrics[w]["in"]["adaptability"][p]
            sh.cell(curr_row, col + 17).value = hist1.metrics[w]["in"]["adaptability"][p]
            sh.cell(curr_row, col + 18).value = hist1.metrics[w]["out"]["adaptability"][p]
            curr_row+=1
    curr_row = 3
    col = 23
    n = hist1.real_horizon
    fh = hist1.fixed_horizon
    
    # calculate and show mean and variance of diffs
    for p in products:
        cqdh1 = getDiffHist(hist1.cproduct_supply, p, n, fh)
        cqdh2 = getDiffHist(hist2.cproduct_supply, p, n, fh)
        mean1, var1 = getMeanVarDiffHist(cqdh1)
        mean2, var2 = getMeanVarDiffHist(cqdh2)
        print("\n", p)
        print("S1 nervousness: ", abs(var1 / mean1))
        print("S2 nervousness: ", abs(var2 / mean2))
        sh.cell(curr_row, col).value = abs(var1 / mean1 )
        sh.cell(curr_row, col + 1).value = abs(var2 / mean2)
        curr_row += 1
        
    wb.save(dst_f)





