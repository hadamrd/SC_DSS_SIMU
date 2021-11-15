import openpyxl

def run(file_path, horizon):
    affiliates = {
        "001": "france",
        "002": "spain",
        "003": "chili",
        "004": "australia"
    }
    supply_demand = {}
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    i = 0
    while sheet.cell(2 + i, 1).value:
        quantity = int(sheet.cell(2 + i, 2).value)
        affiliate = affiliates[sheet.cell(2 + i, 3).value]
        week = int(sheet.cell(2 + i, 6).value.split("/")[0][1:])
        product = sheet.cell(2 + i, 12).value
        if affiliate not in supply_demand:
            supply_demand[affiliate] = {}
        if product not in supply_demand[affiliate]:
            supply_demand[affiliate][product] = [None for _ in range(horizon)]
        supply_demand[affiliate][product][week-2] = quantity
        i += 1
    return supply_demand
    
if __name__ == "__main__":
    run("simu_inputs/supply_plan_S2.xlsx", 24)
