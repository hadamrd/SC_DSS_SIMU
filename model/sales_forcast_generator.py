import math
import random
import openpyxl

affiliates = {
	"france": {"min": 0, "max":200, "products": ["P1", "P2", "P3", "P4"], "code": "001" },
	"spain": {"min": 0, "max":200, "products": ["P1", "P2"], "code": "002" },
	"chili": {"min": 0, "max":300, "products": ["P1", "P3"] , "code": "003"},
	"australia": {"min": 0, "max":400, "products": ["P1", "P2"], "code": "004" }
}

model_wb = openpyxl.load_workbook('sales_forcast_affiliate_model.xlsx')

def accumu(lis):
    total = 0
    for v in lis:
        total += v
        yield total

def randSalesForcast(a, horizon):
    min = affiliates[a]['min']
    max = affiliates[a]['max']
    ans = [10 * random.randint(min//10, max//10) for _ in range(horizon)]
    return ans

def getModelValues(affiliate, t):
    model_sheet = model_wb[affiliates[affiliate]["code"]]
    a = float(model_sheet.cell(row=2, column=t + 2).value)
    b = float(model_sheet.cell(row=3, column=t + 2).value)   
    c = float(model_sheet.cell(row=4, column=t + 2).value) 
    d = float(model_sheet.cell(row=5, column=t + 2).value)
    ref_week = int(model_sheet.cell(row=6, column=t + 2).value)
    return a, b, c, d, ref_week

def pickRand(a, b, c, d):
    alpha = random.random()
    x1 = alpha * (b - a) + a
    x2 = d - alpha * (d - c)
    rd = x1 if random.random() < 0.5 else x2
    rd = round(rd / 10) * 10
    return rd
        
def calcRandCPV(horizon, affiliate, prev_pv):
    prev_cpv = list(accumu(prev_pv))
    min_pv = affiliates[affiliate]['min']
    max_pv = affiliates[affiliate]['max']
    cpv = prev_cpv[1:] + [prev_cpv[-1] + 10 * random.randint(min_pv//10, max_pv//10)]
    
    acpv = [0 for _ in range(horizon)]
    for t in range(horizon):
        a, b, c, d, rw = getModelValues(affiliate, t)
        min_1 = round(cpv[t] + a * (cpv[t] - prev_cpv[rw-1]))
        min_2 = round(cpv[t] + b * (cpv[t] - prev_cpv[rw-1]))
        max_2 = round(cpv[t] + c * (cpv[t] - prev_cpv[rw-1]))
        max_1 = round(cpv[t] + d * (cpv[t] - prev_cpv[rw-1]))
        rd_pv = pickRand(min_1, min_2, max_1, max_2)
        acpv[t] = max(rd_pv, acpv[t-1] if t>0 else 0)
    return acpv
        
def genRandSalesForcast(horizon, affiliate, prev_pv=None):
    if prev_pv is None:
        return randSalesForcast(affiliate, horizon)
    acpv = calcRandCPV(horizon, affiliate, prev_pv)
    new_apv = [None for _ in range(horizon)]
    new_apv[0] = acpv[0] - prev_pv[0]
    for t in range(1, horizon):
        new_apv[t] = acpv[t] - acpv[t-1]
    return new_apv

def run(prev_sales_forcast, horizon):
    return {
        a: {
            p: genRandSalesForcast(horizon, a, prev_sales_forcast[a][p]) for p in aff["products"]
        } for a, aff in affiliates.items()
    }

if __name__=="__main__":
    cm = 0
    cv = 0
    for k in range(10000):
        h = 24
        prev = [random.randint(0, 10000) for _ in range(h)]
        ans = genRandSalesForcast(h, "france", prev)
        prev = prev + [0]
        ans = [0] + ans
        mean = sum([ans[t] - prev[t] for t in range(h+1)]) / (h+1)
        var =  math.sqrt(sum([(ans[t] - prev[t] - mean)**2 for t in range(h+1)]) / (h+1))
        cm += mean
        cv += var
    print("mean diff: ", cm / 10000)
    print("sigma: ", cv / 10000)