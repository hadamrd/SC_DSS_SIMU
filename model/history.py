import openpyxl
import json 
import os
import re
from . import utils, Shared


class History(Shared):
    def __init__(self) -> None:
        super().__init__()
        self.start_week = None
        self.end_week = None
        self.pv  = []
        self.ba  = []
        self.pa  = []
        self.dept = []
        self.pdp = []
        self.bp  = []
        self.s0  = []
        self.product_supply  = []
        self.cproduct_supply = []
        self.metrics = []

    def sumCumHistOverAff(self, cumHist):
        return {p: [[
            sum([
                cumHist[w][a][p][t] for a in self.itProductAff(p)
                ]) for t in range(self.horizon)]
                for w in range(self.nbr_weeks)]
                for p in self.products}

    def init(self, start_week, end_week, with_filter):
        self.with_filter = with_filter
        self.start_week = start_week
        self.end_week = end_week
        self.nbr_weeks = self.end_week - self.start_week + 1

    def fillData(self, snapshot: dict):
        self.ba.append(snapshot["demand"])
        self.pv.append(snapshot["sales_forcast"])
        self.pa.append(snapshot["supply"])
        self.dept.append(snapshot["dept"])
        self.pdp.append(snapshot["reception"])
        self.bp.append(snapshot["prod_demand"])
        self.s0.append(snapshot["initial_stock"])
        self.product_supply.append(snapshot["product_supply"])
        self.cproduct_supply.append(snapshot["cproduct_supply"])
        self.metrics.append(snapshot["metrics"])

    def load(self, history_folder):
        for file_name in os.listdir(history_folder):
            if file_name.startswith("snapshot_S"):
                week = int(re.match(".*S(\d+).*", file_name).group(1))
                start_week = min(week, self.start_week) if self.start_week is not None else week
                end_week = max(week, self.end_week) if self.end_week is not None else week
        self.init(start_week, end_week)
        for file_name in os.listdir(history_folder):
            if file_name.startswith("snapshot_S"):
                with open(os.path.join(history_folder, file_name)) as fp:
                    snapshot = json.load(fp)
                self.fillData(snapshot)

    def exportToExcel(self, prefix, results_folder):
        if not os.path.exists(results_folder):
            os.mkdir(results_folder)

        wb = openpyxl.load_workbook(self.history_template_f)

        for p in self.products:
            dst_file = os.path.join(results_folder, f"{prefix}_{p}_history.xlsx")
            for w in range(self.nbr_weeks):
                for t in range(self.horizon):
                    wb["PV"].cell(row=3+w, column=3+t+w).value = sum([self.pv[w][a][p][t] for a in self.itProductAff(p)])
                    wb["BA"].cell(row=3+w, column=3+t+w).value = sum([self.ba[w][a][p][t] for a in self.itProductAff(p)])
                    wb["BP"].cell(row=3+w, column=3+t+w).value = self.bp[w][p][t]
                    wb["PDP"].cell(row=3+w, column=3+t+w).value = self.pdp[w][p][t]
                    wb["PA"].cell(row=3+w, column=3+t+w).value = self.product_supply[w][p][t]
            wb.save(dst_file)

        wb = openpyxl.load_workbook(self.history_template_f)
        wb.remove_sheet(wb["PDP"])
        wb.remove_sheet(wb["BP"])
        for a in self.itAffiliates():
            for p in self.itAffProducts(a):
                dst_file = os.path.join(results_folder, f"{prefix}_{a}_{p}_history.xlsx")
                for w in range(self.nbr_weeks):
                    for t in range(self.horizon):
                        wb["PV"].cell(row=3+w, column=3+t+w).value = self.pv[w][a][p][t]
                        wb["BA"].cell(row=3+w, column=3+t+w).value = self.ba[w][a][p][t]
                        wb["PA"].cell(row=3+w, column=3+t+w).value = self.pa[w][a][p][t]
                wb.save(dst_file)
