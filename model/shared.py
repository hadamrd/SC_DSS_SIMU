import json
from typing import Any, Tuple
import enum
import math
import openpyxl
from . import utils 
class PvRandStrat(enum.Enum):
    minmax = "minmax"
    uniform = "uniform"
class Shared:
    def __init__(self) -> None:
        with open("config/settings.json") as fp:
            self.settings: dict[str, Any] = json.load(fp)
        self.affiliate: dict                = self.settings["affiliate"]
        self.horizon: int                   = self.settings["horizon"]
        self.prod_time: int                 = self.settings["prod_time"]
        self.real_horizon: int              = self.settings["real_horizon"]
        self.fixed_horizon: int             = self.settings["fixed_horizon"]
        self.l4n_threshold: float           = self.settings["l4n_threshold"]
        self.history_template_f: str        = self.settings["history_template"]
        self.metrics_template_f: str        = self.settings["metrics_template"]
        self.proba_pv_inf: float            = self.settings["proba_pv_inf"]
        self.demand_UCMF: str               = self.settings["demand_UCMF"]
        self.reception_UCMF: str            = self.settings["reception_UCMF"]
        self.sales_UCMF: str                = self.settings["sales_UCMF"]
        self.indicators_template_f: str     = self.settings["indicators_template_f"]
        self.platform_template_f: str       = self.settings["platform_template_f"]
        self.products = set([p for _, p in self.itParams()])
        
    def itAffiliates(self):
        for a in self.affiliate.keys():
            yield a
        
    def getAffCode(self, a) -> str:
        return self.affiliate[a]["code"]
    
    def itAffProducts(self, a):
        return self.affiliate[a]["products"]
    
    def getAffByCode(self, code):
        for a, data in self.affiliate.items():
            if data["code"] == code:
                return a
            
    def getAffDeliveryTime(self, a):
        return self.affiliate[a]["delivery_time"]
    
    def getAffTargetStock(self, a):
        return self.affiliate[a]["target_stock"]
    
    def getAffPvRange(self, a):
        return self.affiliate[a]["pv_range"]
    
    def sumOverAffiliate(self, quantity, product=None, horizon=None) -> list[int]:
        if not horizon:
            horizon = self.horizon
        if product:
            return [sum([quantity[a][product][t] for a in self.itProductAff(product)]) for t in range(horizon)]
        else:
            return {product: [sum([quantity[a][product][t] for a in self.itProductAff(product)]) for t in range(horizon)] for product in self.products}

    def itParams(self):
        for a, data in self.affiliate.items():
            for p in data["products"]:
                yield a, p

    def itProductAff(self, p):
        for a, data in self.affiliate.items():
            if p in data["products"]:
                yield a

    def dipatchSupply(self, capacity, raw_demand, a, p, t):
        if raw_demand[a][p][t] < 0:
            return 0
        tot_raw_demand = sum([raw_demand[a][p][t] for a in self.itProductAff(p) if raw_demand[a][p][t] > 0])
        if capacity[p][t] < 0:
            return 0
        if capacity[p][t] < tot_raw_demand:
            return math.floor(capacity[p][t] * raw_demand[a][p][t] / tot_raw_demand)
        else:
            return raw_demand[a][p][t]

    def dispatch(self, capacity, demand: dict, prev_supply) -> dict[str, dict[str, list[int]]]:
        supply = {a: {p: [None] * self.horizon for p in self.itAffProducts(a)} for a in self.itAffiliates()}   
        for p in self.products:
            for t in range(self.horizon):
                for  a in self.itProductAff(p):
                    if t < self.fixed_horizon:
                        supply[a][p][t] = self.dipatchSupply(capacity, prev_supply, a, p, t)
                        # supply[a][p][t] = prev_supply[a][p][t]
                    else:
                        tot_demand = sum([demand[a][p][t] for a in self.itProductAff(p)])
                        if tot_demand == 0:
                            supply[a][p][t] = round(capacity[p][t] / len(demand.keys()))
                        else:
                            supply[a][p][t] = math.floor(capacity[p][t] * demand[a][p][t] / tot_demand)
        return supply

    def getEmptyAffQ(self, value=None, size=None):
        if not size:
            size = self.horizon
        return {a: {p: [value] * size for p in self.itAffProducts(a)} for a in self.itAffiliates()}
    
    def getEmptyProductQ(self, value=None, size=None):
        if not size:
            size = self.horizon
        return {p: [value] * size for p in self.products}
    
    def loadAffModelFromExcel(self, umcd_f: str, size):
        d_model = {a: {p: {} for p in self.itAffProducts(a)} for a in self.itAffiliates()}
        wb = openpyxl.load_workbook(umcd_f)
        sh = wb.active
        r = 2
        while sh.cell(r, 1).value:
            product = sh.cell(r, 1).value
            aff_code = sh.cell(r, 2).value
            aff = self.getAffByCode(aff_code)
            param = sh.cell(r, 4).value
            if param == "RefWeek":
                quantity = utils.readRefWeekRow(sh, r, 5, size)
            else:
                quantity = utils.readSubRow(sh, r, 5, size)
            d_model[aff][product][param] = quantity
            r += 1
        return d_model

    def loadProductModelFromExcel(self, file_name: str, size) -> None:
        wb = openpyxl.load_workbook(file_name)
        sh = wb.active
        params = ["a", "b", "c", "d", "ModelType", "RefWeek"]
        r_model = {
            p: { 
                param: None for param in params
            } for p in self.products
        }
        r = 2
        while sh.cell(r, 1).value:
            product = sh.cell(r, 1).value
            param = sh.cell(r, 4).value
            if param == "RefWeek":
                quantity = utils.readRefWeekRow(sh, r, 5, size)
            else:
                quantity = utils.readSubRow(sh, r, 5, size)
            r_model[product][param] = quantity
            r += 1
        return r_model