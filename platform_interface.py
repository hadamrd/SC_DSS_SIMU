import openpyxl

def loadSupplyPlan(file_path, affiliate_code, horizon):
    supply_demand = {}
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    i = 0
    while sheet.cell(2 + i, 1).value:
        quantity = int(sheet.cell(2 + i, 2).value)
        affiliate = affiliate_code[sheet.cell(2 + i, 3).value]
        week = int(sheet.cell(2 + i, 6).value.split("/")[0][1:])
        product = sheet.cell(2 + i, 12).value
        if affiliate not in supply_demand:
            supply_demand[affiliate] = {}
        if product not in supply_demand[affiliate]:
            supply_demand[affiliate][product] = [None for _ in range(horizon)]
        supply_demand[affiliate][product][week-2] = quantity
        i += 1
    return supply_demand

def generateInput(file_path, model):
    wb = openpyxl.load_workbook("templates/template_platform_input.xlsx")
    sheet = wb.active
    sheet.cell(2, 2).value = f"W{model.week}/20"
    cdc_prod_plan = model.pa_cdc.getProdPlan()
    cdc_supply_demand = model.pa_cdc.getSupplyDemand()
    cdc_queued_prod = model.pa_cdc.getQueuedProd()
    cdc_initial_stock = model.cbn_cdc.initial_stock
    cdc_prev_supply_plan = model.cdc_supply_plan
    offset = 0
    # header week before
    sheet.cell(4, 8).value = f"W{model.week-1}/20"
    for t in range(model.horizon):
        # header weeks in
        sheet.cell(4, 9 + t).value = f"W{model.week+t}/20"
    for p in model.products:
        product_block_start_row = 5 + offset
        # stock onhand 
        sheet.cell(product_block_start_row + 1, 8).value = cdc_initial_stock[p]
        for t in range(model.horizon):
            # programmed_reception Factory -> CDC (pdp + queued)
            sheet.cell(product_block_start_row, 9 + t).value = cdc_prod_plan[p][t] + cdc_queued_prod[p][t]
            j = 0
            for a in model.affiliates.values():
                if p in a.products:
                    # BA Affiliate -> CDC
                    sheet.cell(product_block_start_row + 2 + j * 2, 9 + t).value = cdc_supply_demand[a.name][p][t]
                    # PA CDC -> affiliate 
                    sheet.cell(product_block_start_row + 2 + j * 2 + 1, 9 + t).value = cdc_prev_supply_plan[a.name][p][t]
                    j += 1
        nbr_ff_p = sum([1 for aff in model.affiliates.values() if p in aff.products])
        offset += (2 * nbr_ff_p + 3)
    wb.save(file_path)
