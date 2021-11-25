from random import randint
import re
import openpyxl
from model.model import Model
import model.utils as utils
import matplotlib.pyplot as plt

class RiskManager:

    def __init__(self, model: Model) -> None:
        self.model = model
        self.horizon = self.model.horizon - 4

    def sumOverAffiliate(self, q, p, param):
        return [sum([q[a][p][param][t] for a in q if p in q[a]]) for t in range(self.horizon)]

    def loadDModel(self, file_name: str):
        """Load demand uncertainty model from excel file

        Args:
            file_name (str): the excel file
        """
        self.d_aff_model = {a: {p: {} for p in aff_products} for a, aff_products in self.model.affiliate_product.items()}
        self.d_model = {p: {} for p in self.model.products}
        wb = openpyxl.load_workbook(file_name)
        sh = wb.active
        r = 2
        while sh.cell(r, 1).value:
            product = sh.cell(r, 1).value
            aff_code = sh.cell(r, 2).value
            aff = self.model.affiliate_code[aff_code]
            param = sh.cell(r, 4).value
            if param == "RefWeek":
                quantity = self.readRefWeekRow(sh, r, 5)
            else:
                quantity = utils.getSubRow(sh, r, 5, self.horizon)
            self.d_aff_model[aff][product][param] = quantity
            r += 1
    
    def getDpm(self, d, p):
        params = ["a", "b", "c", "d"]
        dist = {param: [0] * self.horizon for param in params + ["ref_week", "model_type"]}
        for a in d:
            if p in d[a]:
                Q = list(utils.accumu(d[a][p]))
                for t in range(self.horizon):
                    rw_t = self.d_aff_model[a][p]["RefWeek"][t]
                    model_type = self.d_aff_model[a][p]["ModelType"][t] 
                    for param in params:
                        alpha_t = self.d_aff_model[a][p][param][t]
                        F_t = Q[t] - Q[rw_t-2] if rw_t-2>0 else Q[t]
                        if model_type == "I1":
                            F_t /= t - (rw_t-1) + 1
                        dist[param][t] += Q[t] + alpha_t * F_t
        return dist

    def readRefWeekRow(self, sheet, row, start_col):
        string_ref_weeks = utils.getSubRow(sheet, row, start_col, self.horizon)
        ref_weeks = list(map(int, [re.match(".*W(\d+).*", rw).group(1) for rw in string_ref_weeks]))
        return ref_weeks
        
    def loadRModel(self, file_name: str) -> None:
        wb = openpyxl.load_workbook(file_name)
        sh = wb.active
        params = ["a", "b", "c", "d", "model_type", "ref_week"]
        self.r_model = {p: {
            param: utils.getSubRow(sh, 2 + j + len(params) * k, 5, self.horizon) if param != "ref_week" else 
            self.readRefWeekRow(sh, 2 + j + len(params) * k, 5) 
            for j, param in enumerate(params)
        } for k, p in enumerate(self.model.products) }
    
    def getRpm(self, r:dict, p: str) -> dict:
        params = ["a", "b", "c", "d"]
        Q = list(utils.accumu(r))
        dist = {param: [None] * self.horizon for param in params}
        for t in range(self.horizon):
            rw_t = self.r_model[p]["ref_week"][t]
            model_type = self.r_model[p]["model_type"][t]
            F_t = Q[t] - Q[rw_t-2] if rw_t-2>0 else Q[t]
            for param in params:
                if model_type == "I2":
                    dist[param][t] = round(Q[t] + self.r_model[p][param][t] * F_t)
                elif model_type == "I1":
                    dist[param][t] = round(Q[t] + self.r_model[p][param][t] * F_t / (t - (rw_t-1) + 1))
        return dist

    def tpzd(self, a, b, c, d, x):
        if x < a:
            return 0
        elif x < b:
            return (x - a) / (b - a)
        elif x < c:
            return 1
        elif x < d:
            return (d - x) / (d - c)
        else:
            return 0
    
    def l1p(self, a, b, s0, x):
        if x - s0 < a:
            return 0
        elif x - s0 < b:
            return (x - s0 - a) / (b - a)
        else:
            return 1
    
    def l2p(self, c, d, x):
        if x > d:
            return 0
        elif x > c:
            return (d - x) / (d - c)
        else:
            return 1

    def sampleTrapeze(self, a, b, c, d, nbr_ech):
        min = a - a/20
        max = d + d/20
        ax = utils.linspace(min, max, nbr_ech)
        y = [self.tpzd(a, b, c, d, x) for x in ax]
        return ax, y
    
    def sampleL1Possibility(self, rpm, s0, t, nbr_ech):
        a, b, d = rpm["a"][t], rpm["b"][t], rpm["d"][t]
        min = a - a/20
        max = d + d/20
        ax = utils.linspace(min, max, nbr_ech)
        l1_p = [self.l1p(a, b, s0, x) for x in ax]
        return ax, l1_p
    
    def sampleL2Possibility(self, dpm, t, nbr_ech):
        a, c, d = dpm["a"][t], dpm["c"][t], dpm["d"][t]
        min = a - a/20
        max = d + d/20
        ax = utils.linspace(min, max, nbr_ech)
        l2_p = [self.l2p(c, d, x) for x in ax]
        return ax, l2_p
    
    def sampleL4Possibility(self, dpm, rpm, s0, t, nbr_ech):
        ax_l2, l2_p = self.sampleL2Possibility(dpm, t, nbr_ech)
        ax_l1, l1_p = self.sampleL1Possibility(rpm, s0, t, nbr_ech)
        min_ax = min(ax_l1[0], ax_l2[0])
        max_ax = max(ax_l1[-1], ax_l2[-1])
        ax = utils.linspace(min_ax, max_ax, nbr_ech)
        l4_p = [max(l1, l2) for l2, l1 in zip(l1_p, l2_p)]
        return ax, l4_p

    def getL1Possibility(self, rpm: dict, x: dict, s0: dict) -> dict:
        """calculate L1 possibility for every period 't' in the horizon

        Args:
            fcr (dict): reception possibility model parameters
            x (dict): provisionning plan the cdc is trying to figure out
            s0 (dict): the cdc initial stocks for every product

        Returns:
            dict: L1 possibility
        """
        l1_possibility = [self.l1p(rpm["a"][t], rpm["b"][t], s0, x[t]) for t in range(self.horizon)]
        return l1_possibility

    def getL2Possibility(self, dpm: dict, x: dict) -> dict:
        """calculate L2 possibility for every period 't' in the horizon

        Args:
            dpm (dict): demand possibility model
            x (dict): provisionning plan the cdc is trying to figure out

        Returns:
            dict: L2 possibility
        """
        l2_possibility = [self.l2p(dpm["c"][t], dpm["d"][t], x[t]) for t in range(self.horizon)]
        return l2_possibility

    def getL4Possibility(self, l1p: dict, l2p: dict) -> dict:
        """calculate L4 possibility for every period 't' in the horizon
        Args:
            l1p (dict): L1 possibility for the product 'p'
            l2p (dict): L2 possibility for the product 'p'

        Returns:
            dict: L4 possibility 
        """
        l4_possibility = [max(l1p[t], l2p[t]) for t in range(self.horizon)]
        return l4_possibility
    
    def getBestDeltaX(self, rpm, dpm, s0, x, t):
        b1, c1 = rpm["b"][t], rpm["c"][t]
        b2, c2 = dpm["b"][t], dpm["c"][t]
        alpha = min(c1 + s0, c2)
        chi = max(b1 + s0, b2)
        if alpha < x[t] < chi:
            return min(abs(alpha - x[t]), abs(chi - x[t]))
        else:
            return 0

    
    def getG(self, rpm, dpm, s0, x):
        l1p = self.getL1Possibility(rpm, x, s0)
        l2p = self.getL2Possibility(dpm, x)
        l4p = self.getL4Possibility(l1p, l2p)
        nl4p = [1 - _ for _ in l4p]
        return nl4p, max(nl4p)

    def findSolX(self, rpm, dpm, s0, x):
        g = 1
        nbiter = 0
        while g > 0.5:
            x =  [x[t] + self.getBestDeltaX(rpm, dpm, s0, x, t) for t in range(self.horizon)]
            nl4p, g = self.getG(rpm, dpm, s0, x)
            nbiter += 1
        return x, g




    



if __name__ == "__main__":
    pass
