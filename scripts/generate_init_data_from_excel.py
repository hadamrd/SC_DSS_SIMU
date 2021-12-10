import openpyxl
import json 

def gen(src_file, dst_file, sales_init_f):
    prods = ["P1", "P2", "P3", "P4"]
    aff = {
        "france": ["P1", "P2", "P3", "P4"],
        "spain": ["P1", "P2"],
        "chili": ["P1", "P3"],
        "australia": ["P1", "P2"]
    }
    wb = openpyxl.load_workbook(src_file)
    sh = wb["MJ_Initialisation"]
    h = 24
    

    # get pv
    pv = {a: {p: [None for _ in range(h)] for p in ps} for a, ps in aff.items()}
    offset = 6
    for a, ps in aff.items():
        for p in ps:
            for t in range(h):
                pv[a][p][t] = int(sh.cell(offset, 4+t).value)
            offset+=1
    
    # get pa 
    offset = 20
    pa = {a: {p: [None for _ in range(h)] for p in ps} for a, ps in aff.items()}
    for a, ps in aff.items():
        for p in ps:
            for t in range(h):
                pa[a][p][t] = int(sh.cell(offset, 4+t).value)
            offset+=2

    # get pdp
    offset = 44
    pdp = {p: [None for _ in range(h)] for p in prods}
    for p in prods:
        for t in range(h):
            pdp[p][t] = int(sh.cell(offset, 4+t).value)
        offset+=2
        
    with open(dst_file) as fp:
        data = json.load(fp)

    with open(sales_init_f, 'w') as fp:
        json.dump(pv, fp) 
        
    data["prev_supply"] = pa
    data["prev_production"] = pdp

    with open(dst_file, "w") as fp:
        json.dump(data, fp)
 


if __name__=="__main__":
    sales_init_f = "config/sales_S2.json"
    model_init_f = "config/input_S2.json"
    gen("scripts/s2.xlsm", model_init_f, sales_init_f)
