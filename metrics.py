
import os
import openpyxl
from openpyxl.cell.cell import Cell
from history import History


def periodMean(Q_history: list[list[int]], t: int):
    n = len(Q_history[0])
    res = 0
    for k in range(n):
        res += Q_history[t - k][k]
    return res / n

def periodNervosity(Q_history: list[list[int]], t: int) -> float:
    n = len(Q_history[0])
    res = 0
    for k in range(n-1):
        y = t - k
        res += abs(Q_history[y][k] - Q_history[y-1][k+1])
    return res / (n - 1)

def getQMetric(Q_plan: dict, metric, size: int):
    if type(next(iter(Q_plan.values()))) == dict:
        return {a: {p: [
            metric(Q_plan[a][p], t) for t in range(size - 1, 2 * size - 1)
        ]
        for p in Q_plan[a]} for a in Q_plan}
    elif type(next(iter(Q_plan.values()))) == list:
        return {p: [
            metric(Q_plan[p], t) for t in range(size - 1, 2 * size - 1)
        ]
        for p in Q_plan}
    else:
        raise Exception("Given history is of unknown type: ", type(Q_plan))

def writeRow(sh, row: int, start_col: int, lis: list):
    for t, v in enumerate(lis):
        sh.cell(row, start_col + t).value = v

def generateMetricsResult(hist: History, horizon: int, template_file: str, dst_file: str):
    cum_hist = hist.getCumHistory()
    pa_nervosity = getQMetric(cum_hist.supply_plan, periodNervosity, horizon)
    pdp_nervosity = getQMetric(cum_hist.prod_plan, periodNervosity, horizon)
    bp_nervosity = getQMetric(cum_hist.prod_demand, periodNervosity, horizon)
    ba_nervosity = getQMetric(cum_hist.supply_demand, periodNervosity, horizon)
    pv_nervosity = getQMetric(cum_hist.sales_forcast, periodNervosity, horizon)
    unvailability_mean = getQMetric(hist.unavailability, periodMean, horizon)

    wb = openpyxl.load_workbook(template_file)
    for p in hist.products:
        sheet = wb[p]
        writeRow(sheet, 3, 3, hist.sumOverAffiliate(pv_nervosity, p, horizon))
        writeRow(sheet, 4, 3, hist.sumOverAffiliate(ba_nervosity, p, horizon))
        writeRow(sheet, 5, 3, hist.sumOverAffiliate(pa_nervosity, p, horizon))
        writeRow(sheet, 6, 3, pdp_nervosity[p])
        writeRow(sheet, 7, 3, bp_nervosity[p])

        curr_row = 9
        step = 4
        for idx, a in enumerate(hist.itProductAff(p)):
            sheet.cell(curr_row + idx * step, 1).value = a 
            writeRow(sheet, curr_row + idx, 3, pv_nervosity[a][p])
            writeRow(sheet, curr_row + idx, 3, pv_nervosity[a][p])
            writeRow(sheet, curr_row + idx, 3, pv_nervosity[a][p])
            curr_row += step

        curr_row = 27
        for idx, a in enumerate(hist.itProductAff(p)):
            writeRow(sheet, curr_row + idx, 3, unvailability_mean[a][p])
        
        curr_row = 31
        writeRow(sheet, curr_row + idx, 3, hist.sumOverAffiliate(unvailability_mean, p, horizon))
    wb.save(dst_file)







