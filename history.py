import openpyxl
import json 
import os
import re 
from model import utils


class History:
    def __init__(self, simu_params_f) -> None:
        with open(simu_params_f) as fp:
            self.model_params = json.load(fp)
        self.simu_params_f = simu_params_f
        self.products = self.model_params["products"]
        self.affiliates = self.model_params["affiliates"]
        self.horizon = self.model_params["horizon"]
        self.affiliate_products = self.model_params["affiliate_product"]
        self.start_week = None
        self.end_week = None
        self.sales_forcast = None
        self.prod_plan = None
        self.supply_demand = None
        self.supply_plan = None
        self.prod_demand = None
        self.unavailability = None 

    def sumOverAffiliate(self, quantity, product, horizon):
        return [sum([quantity[a][product][t] for a in self.itProductAff(product)]) for t in range(horizon)]

    def itParams(self):
        for a in self.affiliates:
            for p in self.affiliate_products[a]:
                yield a, p

    def itProductAff(self, p):
        for a in self.affiliates:
            if p in self.affiliate_products[a]:
                yield a

    def getQuantityCumHistory(self, q_history):
        nbr_weeks = len(q_history)
        Q_history = [None] * nbr_weeks
        for w in range(nbr_weeks):
            Qr = Q_history[w-1][0] if w != 0 else 0
            Q_history[w] = list(utils.accumu(q_history[w], Qr))
        return Q_history

    def getCumHistory(self):
        cum_hist = History(self.simu_params_f)
        cum_hist.sales_forcast = {a: {p: self.getQuantityCumHistory(self.sales_forcast[a][p]) for p in self.affiliate_products[a]} for a in self.affiliates}
        cum_hist.supply_demand = {a: {p: self.getQuantityCumHistory(self.supply_demand[a][p]) for p in self.affiliate_products[a]} for a in self.affiliates}
        cum_hist.supply_plan =  {a: {p: self.getQuantityCumHistory(self.supply_plan[a][p]) for p in self.affiliate_products[a]} for a in self.affiliates}
        cum_hist.prod_plan = {p: self.getQuantityCumHistory(self.prod_plan[p]) for p in self.products}
        cum_hist.prod_demand = {p: self.getQuantityCumHistory(self.prod_demand[p]) for p in self.products}
        return cum_hist

    def load(self, history_folder):
        for file_name in os.listdir(history_folder):
            if file_name.startswith("snapshot_S"):
                week = int(re.match(".*S(\d+).*", file_name).group(1))
                self.start_week = min(week, self.start_week) if self.start_week is not None else week
                self.end_week = max(week, self.end_week) if self.end_week is not None else week
        self.nbr_weeks = self.end_week - self.start_week + 1
        
        self.sales_forcast = {a: {p: [None] * self.nbr_weeks for p in self.affiliate_products[a]} for a in self.affiliates}
        self.prod_plan =  {p: [None] * self.nbr_weeks for p in self.products}
        self.supply_demand =  {a: {p: [None] * self.nbr_weeks for p in self.affiliate_products[a]} for a in self.affiliates}
        self.supply_plan =  {a: {p: [None] * self.nbr_weeks for p in self.affiliate_products[a]} for a in self.affiliates}
        self.prod_demand =  {p: [None] * self.nbr_weeks for p in self.products}
        self.unavailability =  {a: {p: [None] * self.nbr_weeks for p in self.affiliate_products[a]} for a in self.affiliates}

        for file_name in os.listdir(history_folder):
            if file_name.startswith("snapshot_S"):
                with open(os.path.join(history_folder, file_name)) as fp:
                    snapshot = json.load(fp)
                w = int(re.match(".*S(\d+).*", file_name).group(1)) - self.start_week
                for p in self.products:
                    for a in self.affiliates:
                        if p in self.affiliate_products[a]:
                            self.supply_demand[a][p][w] = snapshot["supply_demand"][a][p]
                            self.sales_forcast[a][p][w] = snapshot["sales_forcast"][a][p]
                            self.supply_plan[a][p][w] = snapshot["supply_plan"][a][p]
                            self.unavailability[a][p][w] = snapshot["unavailabiliy"][a][p]
                    self.prod_plan[p][w] = snapshot["prod_plan"][p]
                    self.prod_demand[p][w] = snapshot["prod_demand"][p]

    def exportToExcel(self, prefix, results_folder, template_file):
        if not os.path.exists(results_folder):
            os.mkdir(results_folder)

        wb = openpyxl.load_workbook(template_file)
        for p in self.products:
            dst_file = os.path.join(results_folder, f"{prefix}_{p}_results.xlsx")
            for w in range(self.nbr_weeks):
                for t in range(self.horizon):
                    wb["PV"].cell(row=3+w, column=3+t+w).value = sum([self.sales_forcast[a][p][w][t] for a in self.itProductAff(p)])
                    wb["BA"].cell(row=3+w, column=3+t+w).value = sum([self.supply_demand[a][p][w][t] for a in self.itProductAff(p)])
                    wb["BP"].cell(row=3+w, column=3+t+w).value = self.prod_demand[p][w][t]
                    wb["PDP"].cell(row=3+w, column=3+t+w).value = self.prod_plan[p][w][t]
                    wb["PA"].cell(row=3+w, column=3+t+w).value = sum([self.supply_plan[a][p][w][t] for a in self.itProductAff(p)])
            wb.save(dst_file)

        wb = openpyxl.load_workbook(template_file)
        wb.remove_sheet(wb["PDP"])
        wb.remove_sheet(wb["BP"])
        for a in self.affiliates:
            for p in self.affiliate_products[a]:
                dst_file = os.path.join(results_folder, f"{prefix}_{a}_{p}_results.xlsx")
                for w in range(self.nbr_weeks):
                    for t in range(self.horizon):
                        wb["PV"].cell(row=3+w, column=3+t+w).value = self.sales_forcast[a][p][w][t]
                        wb["BA"].cell(row=3+w, column=3+t+w).value = self.supply_demand[a][p][w][t]
                        wb["PA"].cell(row=3+w, column=3+t+w).value = self.supply_plan[a][p][w][t]
                wb.save(dst_file)


if __name__ == "__main__":
    pass
